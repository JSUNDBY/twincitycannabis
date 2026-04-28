#!/usr/bin/env python3
"""
Twin City Cannabis — OCM license-list importer

Reads the official Office of Cannabis Management spreadsheet of licensed
cannabis businesses and turns missing retail-capable shops into stub
entries in scraper/data/manual_dispensaries.json.

The OCM publishes an .xlsx file at:
  https://mn.gov/ocm/public-licensing-data/

The exact filename includes the publication date (e.g.
MN_OCM_licensed_cannabis_businesses_04202026_*.xlsx). Either pass the
URL via --url, or download the file and pass --file.

The importer:
  1. Reads the "Licensed cannabis businesses" sheet.
  2. Filters to retail-capable licenses with a real Retail Site Address.
  3. Skips shops already in the directory (matched by name token overlap
     and city). False positives can be re-added by hand if needed.
  4. Geocodes each new shop's address via OpenStreetMap Nominatim
     (rate-limited to 1 req/sec per their TOS).
  5. Generates a manual_dispensaries.json entry with conservative defaults
     (TCC score 75, no logo, no menu yet, hours blank, region inferred
     from city). Marked "manually_added": true and "ocm_imported": true so
     they're easy to find later for polish.
  6. Writes the merged file back. Idempotent — re-running picks up new
     OCM publications without duplicating existing entries.

Run:
  python3 scraper/ocm_import.py --url https://mn.gov/ocm/assets/<file>.xlsx
  python3 scraper/ocm_import.py --file /path/to/local.xlsx
  python3 scraper/ocm_import.py --dry-run        # preview only

After running, follow up with:
  python3 scraper/update_site.py
  node scripts/build_seo.js
"""

import argparse
import json
import re
import sys
import time
import unicodedata
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

DATA_DIR = Path(__file__).parent / "data"
MANUAL_FILE = DATA_DIR / "manual_dispensaries.json"

METRO_CITIES = {
    "minneapolis", "saint paul", "st paul", "st. paul", "bloomington", "edina",
    "eden prairie", "hopkins", "roseville", "new brighton", "brooklyn park",
    "blaine", "eagan", "burnsville", "woodbury", "lakeville", "rosemount",
    "anoka", "ramsey", "chaska", "jordan", "prior lake", "fridley",
    "mendota heights", "west st paul", "west st. paul", "stillwater",
    "richfield", "cottage grove", "golden valley", "coon rapids",
    "vadnais heights", "little canada", "champlin", "shoreview", "andover",
    "maple grove", "savage", "white bear lake", "brooklyn center",
    "saint anthony", "st anthony", "st. anthony", "robbinsdale", "crystal",
    "new hope",
}


def slug(text):
    text = unicodedata.normalize("NFKD", text or "").encode("ascii", "ignore").decode()
    text = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return text or "unnamed"


def name_tokens(name):
    name = unicodedata.normalize("NFKD", name or "").encode("ascii", "ignore").decode()
    name = re.sub(
        r"\b(llc|inc|company|co|ltd|llp|dispensary|dispensaries|cannabis|the|of|mn)\b",
        "",
        name.lower(),
    )
    return set(t for t in re.split(r"[^a-z0-9]+", name) if len(t) >= 3)


def parse_zip(addr):
    m = re.search(r"\b(\d{5})(?:-\d{4})?\b", addr or "")
    return m.group(1) if m else ""


def gradient_for(name):
    """Deterministic green/teal gradient based on the shop name hash."""
    palette = [
        ("#0c4a6e", "#38bdf8"),
        ("#14532d", "#22c55e"),
        ("#1e3a8a", "#7c3aed"),
        ("#7c2d12", "#f59e0b"),
        ("#0f766e", "#5eead4"),
        ("#581c87", "#c084fc"),
        ("#831843", "#f472b6"),
        ("#365314", "#a3e635"),
    ]
    h = sum(ord(c) for c in name) % len(palette)
    a, b = palette[h]
    return f"linear-gradient(135deg, {a}, {b})"


def initial_for(name):
    parts = [p for p in re.split(r"\s+", name.strip()) if p[:1].isalpha()]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[1][0]).upper()


def geocode(addr):
    """Nominatim geocode with TOS-compliant rate limit."""
    q = urllib.parse.urlencode({"q": addr, "format": "json", "limit": 1})
    url = f"https://nominatim.openstreetmap.org/search?{q}"
    req = urllib.request.Request(url, headers={"User-Agent": "TwinCityCannabis/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read())
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print(f"  geocode error: {e}", file=sys.stderr)
    return None, None


def is_retail_capable(row):
    biz = (row.get("Business Type") or "").strip()
    lic = (row.get("License Type") or "").strip()
    addr = (row.get("Retail Site Address") or "").strip()
    if not addr or addr.upper() == "N/A":
        return False
    return any(
        s in biz or s in lic
        for s in ("Retail", "Retailer", "Microbusiness", "Mezzobusiness", "Medical combination")
    )


def load_existing():
    """Return current manual entries + a name-token index of all TCC shops."""
    manual = json.loads(MANUAL_FILE.read_text()) if MANUAL_FILE.exists() else []
    # Index against TCC.dispensaries in data.js
    data_js = Path(__file__).parent.parent / "js" / "data.js"
    text = data_js.read_text()
    m = re.search(r"TCC\.dispensaries\s*=\s*\[", text)
    start = m.end()
    depth, pos = 1, start
    while depth > 0 and pos < len(text):
        if text[pos] == "[":
            depth += 1
        elif text[pos] == "]":
            depth -= 1
        pos += 1
    block = text[start : pos - 1]
    existing = []
    for m2 in re.finditer(r"id: '([^']+)',\s*\n\s*name: '([^']+)',", block):
        cid, cname = m2.group(1), m2.group(2)
        cm = re.search(rf"id: '{re.escape(cid)}'.*?city: '([^']*)'", block, re.S)
        existing.append({
            "id": cid,
            "name": cname,
            "tokens": name_tokens(cname),
            "city": (cm.group(1) if cm else "").lower(),
        })
    return manual, existing


def already_listed(ocm_name, ocm_city, existing):
    toks = name_tokens(ocm_name)
    city_lower = (ocm_city or "").lower()
    for e in existing:
        if not toks or not e["tokens"]:
            continue
        if e["tokens"] == toks:
            return e["id"]
        # Strong overlap (>=2 tokens) anywhere
        if len(toks & e["tokens"]) >= 2:
            return e["id"]
        # City + 1+ token overlap
        if city_lower and city_lower == e["city"] and toks & e["tokens"]:
            return e["id"]
    return None


def parse_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Licensed cannabis businesses"]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    out = []
    for row in rows:
        d = dict(zip(header, row))
        if not is_retail_capable(d):
            continue
        out.append(d)
    return out


def build_entry(row, lat, lon):
    name = (row.get("D.B.A. (Doing Business As)") or row.get("Legal Business Name") or "").strip()
    addr = (row.get("Retail Site Address") or "").strip()
    city = (row.get("City") or "").strip()
    license_no = (row.get("License Number") or "").strip()
    lic_type = (row.get("License Type") or "").strip()
    biz_type = (row.get("Business Type") or "").strip()

    region = "metro" if city.lower() in METRO_CITIES else "greater-mn"

    features = []
    if "Microbusiness" in lic_type:
        features.append("Microbusiness")
    if "Medical" in lic_type:
        features.append("Medical-licensed")
    if "Mezzobusiness" in lic_type:
        features.append("Mezzobusiness")
    if "Cultivation" in biz_type:
        features.append("Grows their own")
    features.append("Newly licensed")

    return {
        "id": slug(name),
        "name": name,
        "tagline": f"Licensed cannabis dispensary in {city}",
        "address": addr,
        "neighborhood": city,
        "city": city,
        "lat": lat,
        "lng": lon,
        "phone": "",
        "hours": {"weekday": "Check website", "weekend": "Check website",
                  "note": "Hours not yet listed — check the shop's website or call"},
        "website": "",
        "tier": "free",
        "tcc_score": 75,
        "scores": {"pricing": 73, "selection": 73, "service": 75, "lab_testing": 75},
        "review_count": 0,
        "verified": True,
        "features": features,
        "gradient": gradient_for(name),
        "initial": initial_for(name),
        "img": "",
        "manually_added": True,
        "ocm_imported": True,
        "ocm_license": license_no,
        "ocm_license_type": lic_type,
        "region": region,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", help="OCM xlsx URL (downloaded then parsed)")
    ap.add_argument("--file", help="Local path to OCM xlsx (use instead of --url)")
    ap.add_argument("--dry-run", action="store_true", help="Show what would be added; don't write")
    ap.add_argument("--metro-only", action="store_true",
                    help="Only import shops in the Twin Cities metro")
    ap.add_argument("--limit", type=int, default=0,
                    help="Cap number of new shops to import this run (useful for testing)")
    args = ap.parse_args()

    if not args.file and not args.url:
        # Default to the most recently published file
        args.url = "https://mn.gov/ocm/assets/MN_OCM_licensed_cannabis_businesses_04202026_tcm1202-714418.xlsx"

    if args.url and not args.file:
        local = DATA_DIR / "ocm_licensed.xlsx"
        print(f"Downloading {args.url}")
        urllib.request.urlretrieve(args.url, local)
        args.file = str(local)

    rows = parse_xlsx(args.file)
    print(f"Found {len(rows)} retail-capable license rows in OCM file")

    manual, existing = load_existing()
    existing_ids = {e["id"] for e in existing}

    candidates = []
    skipped_existing = 0
    for row in rows:
        name = (row.get("D.B.A. (Doing Business As)") or row.get("Legal Business Name") or "").strip()
        city = (row.get("City") or "").strip()
        if not name:
            continue
        if args.metro_only and city.lower() not in METRO_CITIES:
            continue
        slug_id = slug(name)
        if slug_id in existing_ids:
            skipped_existing += 1
            continue
        match = already_listed(name, city, existing)
        if match:
            skipped_existing += 1
            continue
        candidates.append(row)

    if args.limit and len(candidates) > args.limit:
        print(f"Limiting to first {args.limit} of {len(candidates)} candidates")
        candidates = candidates[: args.limit]

    print(f"Skipped {skipped_existing} already in directory")
    print(f"Will import {len(candidates)} new shops")

    if args.dry_run:
        print("\n=== DRY RUN — preview ===")
        for r in candidates:
            n = (r.get("D.B.A. (Doing Business As)") or r.get("Legal Business Name") or "").strip()
            print(f"  + {n} ({r.get('City')})")
        return

    new_entries = []
    for i, row in enumerate(candidates, 1):
        addr = (row.get("Retail Site Address") or "").strip()
        name = (row.get("D.B.A. (Doing Business As)") or row.get("Legal Business Name") or "").strip()
        print(f"[{i}/{len(candidates)}] geocoding {name} — {addr}")
        lat, lon = geocode(addr)
        time.sleep(1.0)  # Nominatim TOS: 1 req/sec
        if lat is None:
            print(f"  no geocode — skipping {name}")
            continue
        new_entries.append(build_entry(row, lat, lon))

    # Append, dedupe by id
    by_id = {d["id"]: d for d in manual}
    for e in new_entries:
        if e["id"] not in by_id:
            by_id[e["id"]] = e
    final = list(by_id.values())

    MANUAL_FILE.write_text(json.dumps(final, indent=4, ensure_ascii=False) + "\n")
    print(f"\nWrote {len(final)} total manual entries ({len(new_entries)} added this run) to {MANUAL_FILE}")
    print("Next: python3 scraper/update_site.py && node scripts/build_seo.js")


if __name__ == "__main__":
    main()
